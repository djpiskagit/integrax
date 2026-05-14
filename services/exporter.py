"""
services/exporter.py
Экспорт лидов в CSV и XLSX.
"""
import io
import csv
from datetime import datetime

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter


COLUMNS = [
    ("id", "ID"),
    ("display_name", "Пользователь"),
    ("username", "Username"),
    ("message_text", "Сообщение"),
    ("chat_name", "Чат"),
    ("message_date", "Дата"),
    ("niche", "Ниша"),
    ("final_score", "Score"),
    ("intent_score", "Intent"),
    ("activity_score", "Activity"),
    ("niche_score", "Niche"),
    ("matched_keywords", "Ключевые слова"),
]


def export_csv(leads: list[dict]) -> io.BytesIO:
    """Возвращает BytesIO с CSV-данными (UTF-8 BOM для Excel)."""
    output = io.StringIO()
    writer = csv.DictWriter(
        output,
        fieldnames=[col[0] for col in COLUMNS],
        extrasaction="ignore",
        lineterminator="\n",
    )

    # Заголовки на русском
    writer.writerow({col[0]: col[1] for col in COLUMNS})

    for lead in leads:
        writer.writerow({col[0]: lead.get(col[0], "") for col in COLUMNS})

    # UTF-8 BOM — чтобы Excel открывал без кракозябр
    result = io.BytesIO()
    result.write(b"\xef\xbb\xbf")  # BOM
    result.write(output.getvalue().encode("utf-8"))
    result.seek(0)
    return result


def export_xlsx(leads: list[dict]) -> io.BytesIO:
    """Возвращает BytesIO с XLSX-файлом."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Лиды"

    # Стили
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1A1A2E")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    score_fill_high = PatternFill("solid", fgColor="D4EDDA")   # зелёный ≥70
    score_fill_mid = PatternFill("solid", fgColor="FFF3CD")    # жёлтый 40–69
    score_fill_low = PatternFill("solid", fgColor="F8D7DA")    # красный <40

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Заголовки
    headers = [col[1] for col in COLUMNS]
    ws.append(headers)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # Данные
    for row_idx, lead in enumerate(leads, start=2):
        row_data = [lead.get(col[0], "") for col in COLUMNS]
        ws.append(row_data)

        final_score = float(lead.get("final_score", 0))

        for col_idx in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            # Цвет строки по score
            score_col_idx = [c[0] for c in COLUMNS].index("final_score") + 1
            if col_idx == score_col_idx:
                cell.font = Font(bold=True)
                if final_score >= 70:
                    cell.fill = score_fill_high
                elif final_score >= 40:
                    cell.fill = score_fill_mid
                else:
                    cell.fill = score_fill_low

    # Ширина колонок
    col_widths = {
        "id": 6, "display_name": 22, "username": 18,
        "message_text": 50, "chat_name": 20, "message_date": 16,
        "niche": 14, "final_score": 9, "intent_score": 9,
        "activity_score": 9, "niche_score": 9, "matched_keywords": 30,
    }
    for col_idx, (field, _) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(field, 14)

    # Закрепить строку заголовков
    ws.freeze_panes = "A2"

    # Автофильтр
    ws.auto_filter.ref = ws.dimensions

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
